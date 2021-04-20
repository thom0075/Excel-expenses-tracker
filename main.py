#!python3
import openpyxl as xl
import progext
import pyinputplus as pip

workbook = xl.Workbook()
print(workbook.active)
current_sheet = workbook["Sheet"]
# amount_one = float(input("Insert the amount of money spent: "))
# amount_two = float(input("Insert the amount of money spent: "))
amount_one = pip.inputNum(prompt="Insert the amount of money spent: ")
amount_two = pip.inputNum(prompt="Insert the amount of money spent: ")

print("The currently selected sheet is: %s " % (workbook.active))
print(current_sheet)

current_sheet["A1"] = "Money spent: "
current_sheet.column_dimensions['A'].width = 15
progext.nextRow("A1", current_sheet)
print(current_sheet.max_row)
current_sheet["A%s" % (current_sheet.max_row + 1)] = amount_one
current_sheet["A%s" % (current_sheet.max_row + 1)] = amount_two
print("Summing the values...")
progext.sleep(0.5)
current_sheet["A%s" % (current_sheet.max_row + 1)] = "Sum is:"
current_sheet["B%s" % (current_sheet.max_row)] = "=SUM(A2:A3)"
print(f"Sum is {amount_one + amount_two}")
progext.checkIfOpen(workbook)
